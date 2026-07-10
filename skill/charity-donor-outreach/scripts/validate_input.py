"""Validate a donor file against the input schema and the outreach policy.

Usage:
    python validate_input.py --input <donors.csv|.xlsx> --config <campaign.json> [--workdir work]

Outputs (in the work directory):
    validated.csv         rows that passed validation, with computed tier and status
    exceptions.csv        rows that failed, each with specific reasons; no guessing
    corrections.csv       suggested fixes a human can approve and resubmit
    validation_report.json  run summary for dashboards and CI

Exit codes: 0 = ran to completion (exceptions may exist), 2 = bad config or unreadable input.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import donor_rules as rules

VALID_TIER_LABELS = set(rules.FINANCIAL_TIERS) | {"Lapsed", ""}
YES_VALUES = {"yes", "y", "true", "1"}
NO_VALUES = {"no", "n", "false", "0", ""}


def load_config(path: Path) -> tuple[dict, list[str]]:
    errors: list[str] = []
    try:
        config = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return {}, [f"config unreadable: {exc}"]

    for key in ("campaign_type", "as_of_date", "charity_name", "donation_url",
                "signer_name", "signer_title"):
        if not str(config.get(key) or "").strip():
            errors.append(f"config missing required key: {key}")

    if "match_confirmed" not in config:
        errors.append("config missing required key: match_confirmed")
    elif not isinstance(config["match_confirmed"], bool):
        errors.append("match_confirmed must be true or false")
    elif config["match_confirmed"]:
        for key in ("match_sponsor", "match_terms"):
            if not str(config.get(key) or "").strip():
                errors.append(f"match_confirmed is true but {key} is empty")

    campaign = str(config.get("campaign_type") or "")
    if campaign and campaign not in rules.CAMPAIGN_TYPES:
        errors.append(
            f"unknown campaign_type {campaign!r}; expected one of {', '.join(rules.CAMPAIGN_TYPES)}"
        )
    if campaign == "event_fundraiser" and not str(config.get("event_name") or "").strip():
        errors.append("event_fundraiser campaign requires event_name")

    try:
        config["_as_of"] = date.fromisoformat(str(config.get("as_of_date")))
    except (TypeError, ValueError):
        errors.append("as_of_date must be a valid YYYY-MM-DD date")

    return config, errors


def load_rows(path: Path) -> list[dict]:
    if path.suffix.lower() in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("reading Excel files requires openpyxl") from exc
        sheet = load_workbook(path, read_only=True, data_only=True).active
        rows_iter = sheet.iter_rows(values_only=True)
        header = [str(h or "").strip().lower() for h in next(rows_iter)]
        return [
            {header[i]: ("" if cell is None else str(cell)) for i, cell in enumerate(row) if i < len(header)}
            for row in rows_iter
            if any(cell is not None and str(cell).strip() for cell in row)
        ]
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        reader.fieldnames = [(f or "").strip().lower() for f in (reader.fieldnames or [])]
        return [row for row in reader if any((v or "").strip() for v in row.values())]


def validate_rows(
    raw_rows: list[dict], config: dict
) -> tuple[list[dict], list[dict], list[dict]]:
    as_of_year = config["_as_of"].year
    validated: list[dict] = []
    exceptions: list[dict] = []
    corrections: list[dict] = []

    name_counts = Counter(
        (row.get("donor_name") or "").strip().lower() for row in raw_rows
    )

    for index, row in enumerate(raw_rows, start=2):  # row 1 is the header
        errors: list[str] = []
        warnings: list[str] = []
        row_fixes: list[dict] = []
        name = (row.get("donor_name") or "").strip()

        def suggest(field: str, current, value, reason: str) -> None:
            """Record a correction a human can approve; never applied silently."""
            row_fixes.append({
                "row_number": index,
                "donor_name": name or "(missing)",
                "field": field,
                "current_value": "" if current is None else str(current),
                "suggested_value": str(value),
                "reason": reason,
            })

        if not name:
            errors.append("donor_name is missing")
        elif name_counts[name.lower()] > 1:
            errors.append("duplicate donor_name; name is the join key, resolve in the source system")

        gifts: list[tuple[int, float]] = []
        try:
            gifts = rules.parse_gifts(row.get("gifts"))
        except ValueError as exc:
            errors.append(str(exc))

        computed_largest = computed_lifetime = computed_last_year = None
        if gifts:
            computed_largest = max(amount for _, amount in gifts)
            computed_lifetime = sum(amount for _, amount in gifts)
            computed_last_year = max(year for year, _ in gifts)

            future = [year for year, _ in gifts if year > as_of_year]
            if future:
                errors.append(
                    f"gift dated after as_of year {as_of_year}: {sorted(set(future))}"
                )
            if any(year == as_of_year for year, _ in gifts):
                warnings.append(
                    f"gift recorded in the as_of year {as_of_year}, confirm the as_of_date "
                    "is intended (loyalty logic counts only the prior full year)"
                )

            for column, computed in (
                ("largest_gift", computed_largest),
                ("lifetime_total", computed_lifetime),
                ("last_gift_year", computed_last_year),
            ):
                try:
                    stated = rules.parse_money(row.get(column))
                except ValueError as exc:
                    errors.append(f"{column}: {exc}")
                    continue
                if stated is not None and abs(stated - computed) > 0.01:
                    errors.append(
                        f"{column} mismatch: file says {stated:,.0f}, gifts say {computed:,.0f}"
                    )
                    suggest(column, row.get(column), f"{computed:g}",
                            "recomputed from the gift history, which is the source of truth")

        stated_tier = (row.get("tier") or "").strip().title()
        if stated_tier not in VALID_TIER_LABELS:
            errors.append(f"unknown tier label: {row.get('tier')!r}")
            stated_tier = ""

        volunteer_raw = (row.get("volunteer") or "").strip().lower()
        if volunteer_raw in YES_VALUES:
            volunteer = True
        elif volunteer_raw in NO_VALUES:
            volunteer = False
        else:
            errors.append(f"volunteer must be Yes or No, got {row.get('volunteer')!r}")
            volunteer = False

        computed_tier = status = None
        if gifts and computed_lifetime is not None:
            computed_tier = rules.compute_tier(computed_lifetime)
            lapsed = rules.is_lapsed(computed_last_year, as_of_year)
            status = "lapsed" if lapsed else "active"

            if stated_tier == "Lapsed" and not lapsed:
                errors.append(
                    f"file says Lapsed but last gift was {computed_last_year}, "
                    f"within {rules.LAPSED_AFTER_YEARS} years of as_of {as_of_year}"
                )
                suggest("tier", row.get("tier"), computed_tier,
                        "donor is active by date policy; tier recomputed from lifetime giving")
            elif stated_tier in rules.FINANCIAL_TIERS and stated_tier != computed_tier:
                errors.append(
                    f"tier mismatch: file says {stated_tier}, lifetime giving of "
                    f"${computed_lifetime:,.0f} computes to {computed_tier}"
                )
                suggest("tier", row.get("tier"), computed_tier,
                        f"lifetime giving of ${computed_lifetime:,.0f} places this donor in {computed_tier}")
            elif stated_tier in rules.FINANCIAL_TIERS and lapsed:
                warnings.append(
                    f"file states active tier {stated_tier} but donor is lapsed by date policy"
                )

        if errors:
            suggestion_text = "; ".join(
                f"set {fix['field']} to {fix['suggested_value']}" for fix in row_fixes
            )
            exceptions.append({
                "row_number": index,
                "donor_name": name or "(missing)",
                "errors": "; ".join(errors),
                "suggested_correction": suggestion_text or "resolve in the source system",
                "disposition": "excluded from letter generation until a person approves a fix",
            })
            corrections.extend(row_fixes)
            continue

        first_name, last_name = rules.split_name(name)
        validated.append({
            "donor_id": rules.slugify(name),
            "donor_name": name,
            "title": (row.get("title") or "").strip(),
            "first_name": first_name,
            "last_name": last_name,
            "stated_tier": stated_tier,
            "tier": computed_tier,
            "status": status,
            "region": (row.get("region") or "").strip(),
            "gifts": (row.get("gifts") or "").strip(),
            "gift_years": "|".join(str(year) for year, _ in gifts),
            "largest_gift": f"{computed_largest:.2f}",
            "lifetime_total": f"{computed_lifetime:.2f}",
            "last_gift_year": str(computed_last_year),
            "volunteer": "Yes" if volunteer else "No",
            "warnings": " | ".join(warnings),
        })

    return validated, exceptions, corrections


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def run(input_path: Path, config_path: Path, workdir: Path) -> dict:
    config, config_errors = load_config(config_path)
    if config_errors:
        for error in config_errors:
            print(f"CONFIG ERROR: {error}", file=sys.stderr)
        raise SystemExit(2)

    try:
        raw_rows = load_rows(input_path)
    except (OSError, RuntimeError, StopIteration) as exc:
        print(f"INPUT ERROR: {exc}", file=sys.stderr)
        raise SystemExit(2)

    validated, exceptions, corrections = validate_rows(raw_rows, config)

    validated_fields = [
        "donor_id", "donor_name", "title", "first_name", "last_name",
        "stated_tier", "tier", "status", "region", "gifts", "gift_years",
        "largest_gift", "lifetime_total", "last_gift_year", "volunteer", "warnings",
    ]
    write_csv(workdir / "validated.csv", validated, validated_fields)
    write_csv(
        workdir / "exceptions.csv", exceptions,
        ["row_number", "donor_name", "errors", "suggested_correction", "disposition"],
    )
    write_csv(
        workdir / "corrections.csv", corrections,
        ["row_number", "donor_name", "field", "current_value", "suggested_value", "reason"],
    )

    warning_rows = [row for row in validated if row["warnings"]]
    report = {
        "run_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "input_file": str(input_path),
        "as_of_date": str(config.get("as_of_date")),
        "campaign_type": config.get("campaign_type"),
        "rows_in": len(raw_rows),
        "rows_validated": len(validated),
        "rows_excepted": len(exceptions),
        "rows_with_warnings": len(warning_rows),
        "suggested_corrections": len(corrections),
        "exceptions": [
            {"donor_name": e["donor_name"], "errors": e["errors"]} for e in exceptions
        ],
        "warnings": [
            {"donor_name": row["donor_name"], "warnings": row["warnings"]}
            for row in warning_rows
        ],
    }
    (workdir / "validation_report.json").write_text(
        json.dumps(report, indent=2), encoding="utf-8"
    )

    print(f"rows in:            {report['rows_in']}")
    print(f"validated:          {report['rows_validated']}")
    print(f"exceptions:         {report['rows_excepted']}")
    print(f"with warnings:      {report['rows_with_warnings']}")
    print(f"suggested fixes:    {report['suggested_corrections']} (work/corrections.csv, approve then resubmit)")
    for entry in report["exceptions"]:
        print(f"  EXCEPTION  {entry['donor_name']}: {entry['errors']}")
    for entry in report["warnings"]:
        print(f"  WARNING    {entry['donor_name']}: {entry['warnings']}")
    return report


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--config", required=True, type=Path)
    parser.add_argument("--workdir", default=Path("work"), type=Path)
    args = parser.parse_args()
    run(args.input, args.config, args.workdir)


if __name__ == "__main__":
    main()
